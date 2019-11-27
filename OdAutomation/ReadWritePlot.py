
import openpyxl

import numpy
from matplotlib import pyplot as plt


# Give the location of the file
path = "ODnov6new.xlsx"

# workbook object is created
wb_obj = openpyxl.load_workbook(path)

ws1 = wb_obj.get_sheet_by_name("Sheet1")

# Establishing the  instance of object
sheet_obj = wb_obj.active

# seting up the variable
count = 0
m_row = sheet_obj.max_row
avg = []

# finding the variable rows consumed by Plate order 1 will also be same for 2 and 3
for i in range(5, m_row + 1):
    cell_obj = sheet_obj.cell(row=i, column=3)


    if cell_obj.value == None:
        break
    else :
        count=count+1
    # print(cell_obj.value) #printing all the  Plate orders
# print(m_row) #printing final row of plate order
print(count) #printing total rows consumed by Plate order


# multiplying d.f  with raw reading
for i in range(0, count-1 ):
    # Case1Calculation
    DfCase1 = sheet_obj.cell(row=i + 6, column=4)
    RawReadingCase1 = sheet_obj.cell(row=i + 6, column=5)
    CalculationOfCase1 = (2 * DfCase1.value * RawReadingCase1.value)
    # print(CalculationOfCase1)

    # Case2Calculation
    DfCase2 = sheet_obj.cell(row=i + 6, column=10)
    RawReadingCase2 = sheet_obj.cell(row=i + 6, column=11)
    CalculationOfCase2 = (2 * DfCase2.value * RawReadingCase2.value)
    # print(CalculationOfCase2)

    # Case3Calculation
    DfCase3 = sheet_obj.cell(row=i + 6, column=16)
    RawReadingCase3 = sheet_obj.cell(row=i + 6, column=17)
    CalculationOfCase3 = (2 * DfCase3.value * RawReadingCase3.value)
    # print(CalculationOfCase3)

    # Printing the avg of 3 Cases
    avgall = ((CalculationOfCase1 + CalculationOfCase2 + CalculationOfCase3) / 3)
    avg.append(avgall)  # making a lnked lists
print(avg)

Splitavg = [avg[i:i + 3] for i in range(0, len(avg), 3)]
Plotavg = avg
print(Splitavg)
print(Plotavg)
print(len(Splitavg))
lastRowvalue =len(Splitavg[len(Splitavg)-1])
if lastRowvalue == 1:
    Splitavg[len(Splitavg)-1].append(None)
    Splitavg[len(Splitavg) - 1].append(None)

elif lastRowvalue ==2:
    Splitavg[len(Splitavg) - 1].append(None)



print(Splitavg)
j = 0

for row in Splitavg:
    for i in range(0, 3):
        u21 = sheet_obj.cell(row=6 + j, column=21 + 3 * i)
        u21.value = row[i]

    j = j + 1

#Code for datetime
# Python code to convert string to list
TimeDelta = []


#timeStart
def Convert(string):
    li = list(string.split("_"))
    return li

# Driver code
TStart = sheet_obj.cell(row=1, column=2)
t1 = TStart.value
#print(t1)

#print(Convert(t1[0:4]))
#print(t1)
#print("Total years in hrs:")
T1Yearsinhrs =abs(365*24*int(t1[0:4]))
#print(T1Yearsinhrs)
#print("Total months in hrs")
T1months = int(t1[5:7])
#print(T1months)
if  (T1months == 1) or (T1months == 3) or (T1months == 5) or (T1months == 7) or (T1months == 8) or  (T1months == 10)   or (T1months == 12):
    T1monthsinHrs = abs(T1months*31*24)
elif (T1months == 4) or (T1months == 6) or  (T1months == 9) or (T1months == 11):
    T1monthsinHrs = abs(T1months * 30 * 24)
elif (T1months == 2):
    T1monthsinHrs = T1months * 29 * 24
else:
    print("The month range is out of range")
#print(T1monthsinHrs)

T1Days = int(t1[8:10])
#print(T1Days)
T1DayinHrs = T1Days*24
#print(T1DayinHrs)
T1hrs = int(t1[11:13])
#print(T1hrs)
T1min = int(t1[14:16])
T1minInHrs = T1min/60

#print(T1minInHrs)
TotalStartTimeinHrs = T1Yearsinhrs+T1monthsinHrs+T1DayinHrs+T1hrs+T1minInHrs
#print(TotalStartTimeinHrs)


#TimeEnd
for i in range(0,count-1):
    TEndA = sheet_obj.cell(row=6+i, column=3)
    t2A = TEndA.value
    # print(t2)

    # print(Convert(t2[0:4]))
    # print(t2)
    print("Total years in hrs:")
    T2YearsinhrsA = 365 * 24 * int(t2A[0:4])
    print(T2YearsinhrsA)
    print("Total months in hrs")
    T2monthsA = int(t2A[5:7])
    # print(T2months)
    if (T2monthsA == 1) or (T2monthsA == 3) or (T2monthsA == 5) or (T2monthsA == 7) or (T2monthsA == 8) or (
            T2monthsA == 10) or (T2monthsA == 12):
        T2monthsinHrsA = T2monthsA * 31 * 24
    elif (T2monthsA == 4) or (T2monthsA == 6) or (T2monthsA == 9) or (T2monthsA == 11):
        T2monthsinHrsA = T2monthsA * 30 * 24
    elif (T2monthsA == 2):
        T2monthsinHrsA = T2monthsA * 29 * 24
    else:
        print("The month range is out of range")
    print(T2monthsinHrsA)

    T2DaysA = int(t2A[8:10])
    print(T2DaysA)
    T2DayinHrsA = T2DaysA* 24
    # print("Total days in hrs:")
    print(T2DayinHrsA)
    T2hrsA = int(t2A[11:13])
    print("Total hrs in hrs:")
    print(T2hrsA)
    T2minA = float(t2A[14:16])
    # print(T2min)
    T2minInHrsA = float(T2minA / 60)
    # print("Total mins in hrs:")
    # print(T2minInHrs)
    TotalEndTimeinHrsA = T2YearsinhrsA + T2monthsinHrsA + T2DayinHrsA + T2hrsA + T2minInHrsA
    print(TotalEndTimeinHrsA)
    TimeDifA = TotalEndTimeinHrsA - TotalStartTimeinHrs
    #print(TimeDifA)
    TimeDelta.append(TimeDifA)

   #saving The timediff
SplitTimeDelta =  [TimeDelta[i:i + 3] for i in range(0, len(TimeDelta), 3)]
PlotTimeDelta =  TimeDelta
lastRowvalueTimeDelta =len(SplitTimeDelta[len(SplitTimeDelta)-1])
if lastRowvalueTimeDelta == 1:
    SplitTimeDelta[len(SplitTimeDelta)-1].append(None)
    SplitTimeDelta[len(SplitTimeDelta) - 1].append(None)

elif lastRowvalueTimeDelta ==2:
    SplitTimeDelta[len(Splitavg) - 1].append(None)


k=0
for row in SplitTimeDelta:

    for i in range(0, 3):
        u20 = sheet_obj.cell(row=6 + k, column=20 + 3 * i)
        u20.value = row[i]

    k = k + 1


wb_obj.save(path)

#picking up diffrent reactor part to plot



countBB=0
countBS=0
countBO=0
for i in range(5, m_row + 1):
    cell_obj = sheet_obj.cell(row=i, column=20)
    print(countBB)
    if  cell_obj.value == None:
        break
    else:
       countBB = countBB + 1
timebb = []
for i in range(0,countBB-1):
    u20 =  sheet_obj.cell(row=6 +i, column=20)
    timebb.append(u20.value)
print(timebb)


avgBB = []


for i in range(0,countBB-1):
    u20 =  sheet_obj.cell(row=6 +i, column=21)
    avgBB.append(u20.value)
print(avgBB)

for i in range(5, m_row + 1):
    cell_obj = sheet_obj.cell(row=i, column=23)
    print(countBS)
    if  cell_obj.value == None:
        break
    else:
       countBS = countBS + 1
timebs = []
for i in range(0,countBS-1):
    u20 =  sheet_obj.cell(row=6 +i, column=23)
    timebs.append(u20.value)
print(timebs)


avgBS = []
for i in range(0,countBB-1):
    u20 =  sheet_obj.cell(row=6 +i, column=24)
    avgBS.append(u20.value)
print(avgBS)

for i in range(5, m_row + 1):
    cell_obj = sheet_obj.cell(row=i, column=26)
    print(countBO)
    if  cell_obj.value == None:
        break
    else:
       countBO = countBO + 1
timebo = []
for i in range(0,countBO-1):
    u20 =  sheet_obj.cell(row=6 +i, column=26)
    timebo.append(u20.value)
print(timebo)


avgBO = []
for i in range(0,countBO-1):
    u20 =  sheet_obj.cell(row=6 +i, column=27)
    avgBO.append(u20.value)
print(avgBO)



    # print(cell_obj.value) #printing all the  Plate orders
# print(m_row) #printing final row of plate order
plt.figure(1)
plt.subplot(311)
plt.plot(timebo, avgBO, 'ro--' )
plt.ylabel("Avg reading Bo")

plt.subplot(312)
plt.plot(timebs, avgBS, 'ko--')
plt.ylabel("Avg reading Bs")

plt.subplot(313)
plt.plot(timebb, avgBB, 'o--')
plt.ylabel("Avg reading BB")
plt.xlabel("Time dif in hrs")
plt.show()
