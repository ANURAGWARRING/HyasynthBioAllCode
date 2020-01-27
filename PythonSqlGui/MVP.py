import openpyxl
path = "strains.xlsx"

#reading xl
wb_obj = openpyxl.load_workbook(path)
ws1 = wb_obj.get_sheet_by_name("strains")
sheet_obj = wb_obj.active
m_row = sheet_obj.max_row
m_cols = sheet_obj.max_column
countCols = 0
countRows = 0
CheckList = []

m_row = sheet_obj.max_row
m_cols = sheet_obj.max_column
for i in range(1, m_row + 1):
    cell_obj = sheet_obj.cell(row=i, column=1)


    if cell_obj.value == None:
        break
    else :
        countRows=countRows+1
    # print(cell_obj.value) #printing all the  Plate orders
# print(m_row) #printing final row of plate order
print(countRows) #printing total rows consumed by Plate order

for i in range(1, m_cols + 1):
    cell_obj = sheet_obj.cell(row=1, column=i)


    if cell_obj.value == None:
        break
    else :
        countCols=countCols+1

print(countCols)


for i in range(0,countRows):
    for j in range(0, countCols):
        Col1 = sheet_obj.cell(row=i+1, column=1+j)
        CheckList.append(Col1.value)
SplitCheckList = [CheckList[i:i + countCols] for i in range(0, len(CheckList), countCols)]
PopSplitCheckList = SplitCheckList.pop(0)
print(PopSplitCheckList)
print(SplitCheckList)
#Converting 2d list to tuples

ListToTuple = []
for i in SplitCheckList:
    ListToTuple.append(tuple(i))
print(ListToTuple)

#Gui
'''
Here the TreeView widget is configured as a multi-column listbox
with adjustable column width and column-header-click sorting.
'''
try:
    import Tkinter as tk
    import tkFont
    import ttk
except ImportError:  # Python 3
    import tkinter as tk
    import tkinter.font as tkFont
    import tkinter.ttk as ttk

class MultiColumnListbox(object):
    """use a ttk.TreeView as a multicolumn ListBox"""

    def __init__(self):
        self.tree = None
        self._setup_widgets()
        self._build_tree()

    def _setup_widgets(self):
        s = """Strains
        """
        msg = ttk.Label(wraplength="4i", justify="left", anchor="n",
            padding=(10, 2, 10, 6), text=s)
        msg.pack(fill='x')
        container = ttk.Frame()
        container.pack(fill='both', expand=True)
        # create a treeview with dual scrollbars
        self.tree = ttk.Treeview(columns=car_header, show="headings")
        vsb = ttk.Scrollbar(orient="vertical",
            command=self.tree.yview)
        hsb = ttk.Scrollbar(orient="horizontal",
            command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set,
            xscrollcommand=hsb.set)
        self.tree.grid(column=0, row=0, sticky='nsew', in_=container)
        vsb.grid(column=1, row=0, sticky='ns', in_=container)
        hsb.grid(column=0, row=1, sticky='ew', in_=container)
        container.grid_columnconfigure(0, weight=1)
        container.grid_rowconfigure(0, weight=1)

    def _build_tree(self):
        for col in car_header:
            self.tree.heading(col, text=col.title(),
                command=lambda c=col: sortby(self.tree, c, 0))
            # adjust the column's width to the header string
            self.tree.column(col,
                width=tkFont.Font().measure(col.title()))

        for item in car_list:
            self.tree.insert('', 'end', values=item)
            # adjust column's width if necessary to fit each value
            for ix, val in enumerate(item):
                col_w = tkFont.Font().measure(val)
                if self.tree.column(car_header[ix],width=None)<col_w:
                    self.tree.column(car_header[ix], width=col_w)

def sortby(tree, col, descending):
    """sort tree contents when a column header is clicked on"""
    # grab values to sort
    data = [(tree.set(child, col), child) \
        for child in tree.get_children('')]
    # if the data to be sorted is numeric change to float
    #data =  change_numeric(data)
    # now sort the data in place
    data.sort(reverse=descending)
    for ix, item in enumerate(data):
        tree.move(item[1], '', ix)
    # switch the heading so it will sort in the opposite direction
    tree.heading(col, command=lambda col=col: sortby(tree, col, \
        int(not descending)))

# the test data ...

car_header = PopSplitCheckList
car_list = SplitCheckList



if __name__ == '__main__':
    root = tk.Tk()
    root.title("Strains")
    listbox = MultiColumnListbox()
    root.mainloop()
