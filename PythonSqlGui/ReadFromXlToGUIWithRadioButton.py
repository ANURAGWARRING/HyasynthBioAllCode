import tkinter as tki # tkinter in Python 3
root = tki.Tk()
frm = tki.Frame(root, bd=16)
frm.grid()

import openpyxl
path = "strains.xlsx"
var = tki.StringVar()
wb_obj = openpyxl.load_workbook(path)
ws1 = wb_obj.get_sheet_by_name("strains")
sheet_obj = wb_obj.active
countCols = 0
countRows = 0
CheckList = []

m_row = sheet_obj.max_row
m_cols = sheet_obj.max_column
for i in range(1, m_row + 1):
    cell_obj = sheet_obj.cell(row=i, column=1)


    if cell_obj.value == None:
        break
    else :
        countRows=countRows+1
    # print(cell_obj.value) #printing all the  Plate orders
# print(m_row) #printing final row of plate order
print(countRows) #printing total rows consumed by Plate order

for i in range(1, m_cols + 1):
    cell_obj = sheet_obj.cell(row=1, column=i)


    if cell_obj.value == None:
        break
    else :
        countCols=countCols+1

print(countCols)

SplitCheckList = [CheckList[i:i + l] for i in range(0, len(CheckList), 2)]
for i in range(0,countRows-1):
    for j in range(0, countCols-1):
        Col1 = sheet_obj.cell(row=i+1, column=1+j)
        CheckList.append(Col1.value)
SplitCheckList = [CheckList[i:i + countCols] for i in range(0, len(CheckList), countCols)]
print(SplitCheckList)

for i in range(len(SplitCheckList)):
    j = 0
    for column in SplitCheckList[i]:
        print(column)
        mild = tki.Radiobutton(frm, text=column, variable=var,bg='SlateBlue3', fg='white' )
        mild.config(indicatoron=0, bd=4, width=12, value='Mild')
        mild.grid(row=i, column=j)
        j=j+1

root.mainloop()
