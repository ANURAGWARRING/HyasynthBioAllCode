#Setting up class Strains,Defining attributes,Clling it's instance

class Strains:
    def __init__(self):
        """ Create a new point at the origin """
        self.id = []
        self.cannabis_controlled_strain = []
        self.narcotic_controlled_strain = []
        self.user_id = []
        self.box = []
        self.slot = []
        self.stocked_in_triplicate =[]
        self.parent_strain_id = []
        self.piCAS_cured_status =[]
        self.organism_id = []
        self.media =[]
        self.selection=[]
        self.temperature=[]
        self.date_created=[]
        self.comments= []
        self.tracking_id=[]
        self.strain_number=[]


#Reading data from xl and converting them to 2d lists

import openpyxl
path = "strains.xlsx"
wb_obj = openpyxl.load_workbook(path)
ws1 = wb_obj.get_sheet_by_name("strains")
sheet_obj = wb_obj.active
m_row = sheet_obj.max_row
m_cols = sheet_obj.max_column
countCols = 0
countRows = 0
m_row = sheet_obj.max_row
m_cols = sheet_obj.max_column
WholeXlData = []
for i in range(1, m_row + 1): #Finding the last column in the xl
    cell_obj = sheet_obj.cell(row=i, column=1)


    if cell_obj.value == None:
        break
    else :
        countRows=countRows+1

for i in range(1, m_cols + 1): #Finding the last row in the xl
    cell_obj = sheet_obj.cell(row=1, column=i)


    if cell_obj.value == None:
        break
    else :
        countCols=countCols+1
for i in range(0,countCols): #adding whole xl data to lists
    for j in range(1, countRows):
        Col1 = sheet_obj.cell(row=j+1, column=1+i)
        WholeXlData.append(Col1.value)

SplitWholeXlData = [WholeXlData[i:i + (countRows-1)] for i in range(0, len(WholeXlData), (countRows-1))] #Converting the 1d list to 2d list
print(SplitWholeXlData)
#appending list values to attributes
Instance1Strains = Strains()
setattr(Instance1Strains, 'id', SplitWholeXlData[0])
setattr(Instance1Strains, 'cannabis_controlled_strain',SplitWholeXlData[1] )
setattr(Instance1Strains, 'narcotic_controlled_strain', SplitWholeXlData[2])
setattr(Instance1Strains, 'user_id', SplitWholeXlData[3])
setattr(Instance1Strains, 'box', SplitWholeXlData[4])
setattr(Instance1Strains, 'slot', SplitWholeXlData[5])

setattr(Instance1Strains, 'stocked_in_triplicate', SplitWholeXlData[6])
setattr(Instance1Strains, 'parent_strain_id', SplitWholeXlData[7])
setattr(Instance1Strains, 'piCAS_cured_status', SplitWholeXlData[8])
setattr(Instance1Strains, 'organism_id', SplitWholeXlData[9])
setattr(Instance1Strains, 'media', SplitWholeXlData[10])
setattr(Instance1Strains, 'selection', SplitWholeXlData[11])
setattr(Instance1Strains, 'temperature', SplitWholeXlData[12])
setattr(Instance1Strains, 'date_created', SplitWholeXlData[13])
setattr(Instance1Strains, 'comments', SplitWholeXlData[14])
setattr(Instance1Strains, 'tracking_id', SplitWholeXlData[15])
setattr(Instance1Strains, 'strain_number', SplitWholeXlData[16])
print(tuple(Instance1Strains.id))
#Setting up  gui

try:
    import Tkinter as tk
    import tkFont
    import ttk
except ImportError:  # Python 3
    import tkinter as tk
    import tkinter.font as tkFont
    import tkinter.ttk as ttk

class MultiColumnListbox(object):
    """use a ttk.TreeView as a multicolumn ListBox"""

    def __init__(self):
        self.tree = None
        self._setup_widgets()
        self._build_tree()

    def _setup_widgets(self):
        s = """Strains
        """
        msg = ttk.Label(wraplength="4i", justify="left", anchor="n",
            padding=(10, 2, 10, 6), text=s)
        msg.pack(fill='x')
        container = ttk.Frame()
        container.pack(fill='both', expand=True)
        # create a treeview with dual scrollbars
        self.tree = ttk.Treeview(columns=Strain_header, show="headings")
        vsb = ttk.Scrollbar(orient="vertical",
            command=self.tree.yview)
        hsb = ttk.Scrollbar(orient="horizontal",
            command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set,
            xscrollcommand=hsb.set)
        self.tree.grid(column=0, row=0, sticky='nsew', in_=container)
        vsb.grid(column=1, row=0, sticky='ns', in_=container)
        hsb.grid(column=0, row=1, sticky='ew', in_=container)
        container.grid_columnconfigure(0, weight=1)
        container.grid_rowconfigure(0, weight=1)

    def _build_tree(self):
        for col in Strain_header:
            self.tree.heading(col, text=col.title(),
                command=lambda c=col: sortby(self.tree, c, 0))
            # adjust the column's width to the header string
            self.tree.column(col,
                width=tkFont.Font().measure(col.title()))

        for item in Strain_list:
            self.tree.insert('', 'end', values=item)
            # adjust column's width if necessary to fit each value
            for ix, val in enumerate(item):
                col_w = tkFont.Font().measure(val)
                if self.tree.column(Strain_header[ix],width=None)<col_w:
                    self.tree.column(Strain_header[ix], width=col_w)

def sortby(tree, col, descending):
    """sort tree contents when a column header is clicked on"""
    # grab values to sort
    data = [(tree.set(child, col), child) \
        for child in tree.get_children('')]
    # if the data to be sorted is numeric change to float
    #data =  change_numeric(data)
    # now sort the data in place
    data.sort(reverse=descending)
    for ix, item in enumerate(data):
        tree.move(item[1], '', ix)
    # switch the heading so it will sort in the opposite direction
    tree.heading(col, command=lambda col=col: sortby(tree, col, \
        int(not descending)))

#Displaying the attributes as gui

Strain_header = ['id','cannabis_controlled_strain','narcotic_controlled_strain','user_id','box','slot','stocked_in_triplicate','parent_strain_id','piCAS_cured_status',
                 'organism_id','media','selection','temprature','date_created','comments','tracking_id','strain_number']
Strain_list = [[Instance1Strains.id[i+0],
                Instance1Strains.cannabis_controlled_strain[i+0],
                Instance1Strains.narcotic_controlled_strain[i+0],
                Instance1Strains.user_id[i+0],
                Instance1Strains.box[i+0],
                Instance1Strains.slot[i+0],
                Instance1Strains.stocked_in_triplicate[i+0],
                Instance1Strains.parent_strain_id[i+0],
                Instance1Strains.piCAS_cured_status[i+0],
                Instance1Strains.organism_id[i+0],
                Instance1Strains.media[i+0],
                Instance1Strains.selection[i+0],
                Instance1Strains.temperature[i+0],
                Instance1Strains.date_created[i+0],
                Instance1Strains.comments[i+0],
                Instance1Strains.tracking_id[i+0],Instance1Strains.strain_number[i+0]]for i in range(countRows-1)]

if __name__ == '__main__':
    root = tk.Tk()
    root.title("Strains")
    listbox = MultiColumnListbox()
    root.mainloop()