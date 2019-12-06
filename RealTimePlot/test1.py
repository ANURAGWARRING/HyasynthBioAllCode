import time
import serial
import re
from matplotlib import pyplot as plt
import numpy as np
from matplotlib import style
import numpy
import openpyxl
from openpyxl import Workbook

# set up the serial line
ser = serial.Serial('COM8', 9600)
print(ser)
time.sleep(3)


CO2Final = []
TimeFinal = []
A0_A4V_Final =  []
A1_A5V_Final = []
A2_A6V_Final =  []
Sensor_A0_A4V_Final =  []
Sensor_A1_A5V_Final = []
Sensor_A2_A6V_Final =  []


for i in range(5):
     b = ser.readline()         # read a byte string
     string_n = b.decode()
     FullRow = string_n.rstrip() # remove \n and \r
     SplitFullRow = re.split(',', FullRow)
     #Getting each variable of list
     TotalTime = float(SplitFullRow[1])
     CO2 = float(SplitFullRow[3])
     A0_A4V = float(SplitFullRow[5])
     A1_A5V = float(SplitFullRow[7])
     A2_A6V = float(SplitFullRow[9])
     Sensor_A0_A4V = float(SplitFullRow[11])
     Sensor_A1_A5V = float(SplitFullRow[13])
     Sensor_A2_A6V = float(SplitFullRow[15])
     #adding each value to dynamic array
     TimeFinal.append(TotalTime)
     CO2Final.append(CO2)
     A0_A4V_Final.append(A0_A4V)
     A1_A5V_Final.append(A1_A5V)
     A2_A6V_Final.append(A2_A6V)
     Sensor_A0_A4V_Final.append(Sensor_A0_A4V)
     Sensor_A1_A5V_Final.append(Sensor_A1_A5V)
     Sensor_A2_A6V_Final.append(Sensor_A2_A6V)
     print( SplitFullRow)
     plt.figure(1)
     plt.title("Voltage vs time in Seconds")
     plt.subplot(311)
     plt.plot(TotalTime, A0_A4V, 'ro--')
     plt.ylabel("Sensor A0_A4v")
     plt.subplot(312)
     plt.plot(TotalTime, A1_A5V, 'ko--')
     plt.ylabel("Sensor A1_A5v")
     plt.subplot(313)
     plt.plot(TotalTime, A2_A6V, 'o--')
     plt.ylabel("Sensor A2_A6v")
     plt.xlabel("Time dif in hrs")
     plt.pause(5)
     plt.figure(2)
     plt.title("Sensor_values vs time in Seconds")
     plt.subplot(311)
     plt.plot(TotalTime, Sensor_A0_A4V, 'ro--')
     plt.ylabel("Voltage A0_A4v")
     plt.subplot(312)
     plt.plot(TotalTime, Sensor_A1_A5V, 'ko--')
     plt.ylabel("Voltage A1_A5v")
     plt.subplot(313)
     plt.plot(TotalTime, Sensor_A2_A6V, 'o--')
     plt.ylabel("Voltage A2_A6v")
     plt.xlabel("Time dif in hrs")
     plt.pause(5)
plt.show()

print(CO2Final)
print(TimeFinal)
#saving the workbook
wb = Workbook()
ws =  wb.active
ws.title = "Changed Sheet"
ws['A1'] = "Time"
ws['B1'] = "Co2"
ws['C1'] = "CO2-Value"
ws['D1'] = "A0/A4V"
ws['E1'] = "A0/A4V-value"
ws['F1'] = "A1/A5V"
ws['G1'] = "A1/A5V-value"
ws['H1'] = "A2/A6V"
ws['I1'] = "A2/A6V-value"
ws['J1'] = "Sensor-A0/A4V"
ws['K1'] = "Sensor-A0/A4V-value"
ws['L1'] = "Sensor-A1/A5V"
ws['M1'] = "Sensor-A1/A5V-value"
ws['N1'] = "Sensor-A2/A6V"

ws['O1'] = "Sensor-A2/A6V-value"

for i in range(5):
     ws.cell(row=i+2, column=1).value = TimeFinal[i]
     ws.cell(row=i+2, column=3).value = CO2Final[i]
     ws.cell(row=i + 2, column=5).value = A0_A4V_Final[i]
     ws.cell(row=i + 2, column=7).value = A1_A5V_Final[i]
     ws.cell(row=i + 2, column=9).value = A2_A6V_Final[i]
     ws.cell(row=i + 2, column=11).value = Sensor_A0_A4V_Final[i]
     ws.cell(row=i + 2, column=13).value = Sensor_A1_A5V_Final[i]
     ws.cell(row=i + 2, column=15).value = Sensor_A2_A6V_Final[i]

val = input("what should be the name of file: ")
print(val)

wb.save(filename =  val)




ser.close()


