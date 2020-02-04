import openpyxl
import math

def ALLCal(ReadCol,WriteColAvg,WriteColSd):
    path = "Example for data processing (3).xlsx"
    wb_obj = openpyxl.load_workbook(path)
    sheet_obj = wb_obj.active
    count = 0
    NumOfSamples = sheet_obj.cell(row=1, column=5).value
    IntNumOfSamples = int(NumOfSamples)
    print("Total number of samples:", end='')
    print(IntNumOfSamples)
    SumOfSamples = []
    AvgOfSamplesR = []
    OrganSamples = []
    NewOrganSamples = []

    Sd = []
    Dilution = sheet_obj.cell(row=3, column=5).value
    NumberOfReplicates = sheet_obj.cell(row=2, column=5).value
    IntNumberOfReplicates = int(NumberOfReplicates)
    print("Total number of Replicates:", end='')
    print(IntNumberOfReplicates)
    for i in range(1, IntNumOfSamples + 1):
        SumOfSamples.append(0)
    print(SumOfSamples)
    m_row = sheet_obj.max_row
    for i in range(10, m_row + 1):
        cell_obj = sheet_obj.cell(row=i, column=4)
        count = count + 1
        if cell_obj.value == None:
            break
    print("Total Occupied Rows:", end='')
    print(count)
    for j in range(0, count - 1):
        cell_obj = sheet_obj.cell(row=10 + j, column=ReadCol)
        OrganSamples.append(cell_obj.value)
        print(NewOrganSamples)
    for j in range(0, IntNumOfSamples):
        for i in range(0, count - 1):
            if ((IntNumOfSamples * i + j) < (count - 1)):
                NewOrganSamples.append(OrganSamples[IntNumOfSamples * i + j])
    for j in range(0, count - 1):
        if (NewOrganSamples[j] == None):
            NewOrganSamples[j] = 0
    print("Total number of NewOrganSamples:", end='')
    print(NewOrganSamples)
    SplitNewOrganOfSamples = [NewOrganSamples[i:i + IntNumberOfReplicates] for i in
                              range(0, len(NewOrganSamples), IntNumberOfReplicates)]
    print("Total number of SplitNewOrganOfSamples:", end='')
    print(SplitNewOrganOfSamples)
    for row in SplitNewOrganOfSamples:
        NoneCount = 0
        k = 0
        SumOfSamples[k] = 0
        for i in row:
            if (i == 0):
                NoneCount = NoneCount + 1
            SumOfSamples[k] = i + SumOfSamples[k]
        print("SumOfSamples:", end='')
        print(SumOfSamples[k])
        print("NoneCount:", end='')
        print(NoneCount)
        if ((IntNumberOfReplicates - NoneCount) == 0):
            AvgOfSamplesR.append(0)
        else:
            AvgOfSamplesR.append((SumOfSamples[k]) / (IntNumberOfReplicates - NoneCount))

    print(AvgOfSamplesR)

    for i in range(0, len(AvgOfSamplesR)):
        cell_obj = sheet_obj.cell(row=i + 10, column=WriteColAvg)
        cell_obj.value = AvgOfSamplesR[i]*Dilution

    #standard Deviation
    for i in range(0,len(SplitNewOrganOfSamples)):
        print(SplitNewOrganOfSamples[i])
        Sum1 = 0
        for j in SplitNewOrganOfSamples[i]:
            print(j)

            Sum1 = (AvgOfSamplesR[i]-j)**2+Sum1
        Sd.append(math.sqrt(Sum1/3))

    print(Sd)
    for i in range(0, len(AvgOfSamplesR)):
        cell_obj = sheet_obj.cell(row=i + 10, column=WriteColSd)
        cell_obj.value = AvgOfSamplesR[i]*Dilution

    wb_obj.save(path)




ALLCal(5,15,25)
ALLCal(6,16,26)
ALLCal(7,17,27)
ALLCal(8,18,28)
ALLCal(9,19,29)
ALLCal(10,20,30)
ALLCal(11,21,31)
ALLCal(12,22,32)

