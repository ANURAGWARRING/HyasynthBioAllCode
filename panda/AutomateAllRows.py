import openpyxl 
  
# Give the location of the file 
path = "Example for data processing (3).xlsx"
  
# workbook object is created 
wb_obj = openpyxl.load_workbook(path) 
  
sheet_obj = wb_obj.active
count=0
count1=0
countin = 0
countout  = 0
x=10
m_row = sheet_obj.max_row 

#finding avg for variable 
for i in range(10, m_row + 1): 
    cell_obj = sheet_obj.cell(row = i, column = 4)
    count=count+1
    
    if cell_obj.value == None:
        break
    #print(cell_obj.value)
#print(m_row)
#print(count)
for i in range(10,count+9):
    cell_obj = sheet_obj.cell(row = i, column = 4)
    if(cell_obj.value != None):
        count1=count1+cell_obj.value;
print(count1)
#print(count1/count)




#finding avg for fixed number of datapoints in the column:
ws = wb_obj.active
Cell_selected = ws['E2']
Value = int(Cell_selected.value)
#print(Cell_selected)
#print(Value)


for i in range (10,count+9):
    countin=0
    #print(x)
    for  i in range (x,x+Value):
        cell_obj = sheet_obj.cell(row = i, column = 4)
        if(cell_obj.value != None):
            countin = countin + int(cell_obj.value)
            x=x+1
    print(countin)
        
    
    
    
    
    

    
    
    
    
    
