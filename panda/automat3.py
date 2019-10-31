# import load_workbook
from openpyxl import load_workbook
# set file path
filepath="Example for data processing (3).xlsx"
# load demo.xlsx 
wb=load_workbook(filepath)
# select demo.xlsx
sheet=wb.active
# get b1 cell value
b1=sheet['B1']
# get b2 cell value
b2=sheet['B2']
# get b3 cell value
b3=sheet.cell(row=3,column=2)
# print b1, b2 and b3
print(b1.value)
print(b2.value)
print(b3.value)