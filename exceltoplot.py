import openpyxl
import numpy
from matplotlib import pyplot as plt
 #Give the location of the file
path = r'C:\Users\warri\PycharmProjects\XltoPLots\Example for data processing (3).xlsx'

# workbook object is created
wb_obj = openpyxl.load_workbook(path)

ws1 = wb_obj.get_sheet_by_name("Sheet2")
sheet_obj = wb_obj.active

timestamp = []
A0V = []
A1V = []
A2V = []
A0VSensor = []
A1VSensor = []
A2VSensor = []


count = 0
count1 = 0

m_row = sheet_obj.max_row

for i in range(1, m_row + 1):
    cell_obj = sheet_obj.cell(row=i*30, column=2)


    if cell_obj.value == None:
        break
    else :
        count=count+1
print(count)

for i in range(720, count-1):
    EachTime = sheet_obj.cell(row=i*30, column=2)
    timestamp.append((EachTime.value/3600))

    EachA0V =  sheet_obj.cell(row=i*30, column=6)
    NewvalueA0v = float (EachA0V.value/1.5272)**((-1)/0.277)
    A0V.append(EachA0V.value)
    A0VSensor.append(NewvalueA0v)

    EachA1V = sheet_obj.cell(row=i * 30, column=8)
    NewvalueA1v =  float (EachA1V.value/0.836)**((1)/(-0.318))
    A1V.append(EachA1V.value)
    A1VSensor.append(NewvalueA1v)

    EachA2V = sheet_obj.cell(row=i * 30, column=10)
    NewvalueA2v =  (0.083)*(EachA2V.value/2.89)**((1)/(-0.164))
    A2V.append(EachA2V.value)
    A2VSensor.append(NewvalueA2v)



#timestamp[0:2] = []
#A0V[0:2] = []
#A1V[0:2] = []
#A2V[0:2] = []
#A0V[3:13]= [0.02,0.02,0.02,0.02,0.02,0.02,0.02,0.02,0.02,0.02]
#A0V[51]= 0.02
print(timestamp)
print(A0V)
print(A1V)
print(A2V)



#print(A0V)




plt.figure(1)
plt.subplot(311)
plt.plot(timestamp, A0V, 'o--' )
plt.ylabel("Voltage BB")

plt.subplot(312)
plt.plot(timestamp, A1V, 'ko--')
plt.ylabel("Voltage Bs")

plt.subplot(313)
plt.plot(timestamp, A2V, 'ro--')
plt.ylabel("Voltage BO")
plt.xlabel("Time dif in hrs")

plt.figure(2)
plt.subplot(311)
plt.plot(timestamp, A0VSensor, 'o--' )
plt.ylabel("Sensor Etoh  BB")

plt.subplot(312)
plt.plot(timestamp, A1VSensor, 'ko--')
plt.ylabel("Sensor Etoh  BS")

plt.subplot(313)
plt.plot(timestamp, A2VSensor, 'ro--')
plt.ylabel("Sensor Etoh  BO")
plt.xlabel("Time dif in hrs")
plt.show()
